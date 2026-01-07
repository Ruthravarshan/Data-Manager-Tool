"""
Helper script to create sample Excel files for testing
Run this to generate test data in data_source folder
"""
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill

def create_sample_excel_files():
    """Create sample Excel files with test data"""
    
    # Create data_source folder if it doesn't exist
    folder_path = "./data_source"
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
        print(f"✓ Created {folder_path} folder")
    
    # Define sample files to create
    # Format: (filename, sheet_name, headers, sample_rows)
    files_to_create = [
        (
            "dm_raw_20251223_100000.xlsx",
            "Demographics",
            ["Subject ID", "Age", "Sex", "Race", "Country"],
            [
                ["100-001", 45, "Male", "White", "USA"],
                ["100-002", 52, "Female", "Black", "USA"],
                ["100-003", 61, "Male", "Asian", "USA"],
            ]
        ),
        (
            "ae_raw_20251223_100100.xlsx",
            "Adverse Events",
            ["Subject ID", "Event", "Onset Date", "Severity", "Outcome"],
            [
                ["100-001", "Headache", "2025-01-15", "Mild", "Resolved"],
                ["100-002", "Nausea", "2025-01-20", "Moderate", "Ongoing"],
            ]
        ),
        (
            "sv_raw_20251223_100200.xlsx",
            "Subject Visits",
            ["Subject ID", "Visit", "Visit Date", "Investigator"],
            [
                ["100-001", "Visit 1", "2025-01-10", "Dr. Smith"],
                ["100-002", "Visit 1", "2025-01-11", "Dr. Johnson"],
                ["100-003", "Visit 2", "2025-01-25", "Dr. Smith"],
            ]
        ),
        (
            "lb_raw_20251223_100300.xlsx",
            "Laboratory",
            ["Subject ID", "Test Name", "Result", "Unit", "Normal Range"],
            [
                ["100-001", "WBC", "7.2", "10^3/µL", "4.5-11.0"],
                ["100-002", "HGB", "13.5", "g/dL", "12.0-16.0"],
            ]
        ),
    ]
    
    # Create each file
    for filename, sheet_name, headers, rows in files_to_create:
        file_path = os.path.join(folder_path, filename)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Add headers with formatting
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Add data rows
        for row_num, row_data in enumerate(rows, 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Adjust column widths
        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15
        
        # Save file
        wb.save(file_path)
        file_size = os.path.getsize(file_path)
        print(f"✓ Created {filename} ({file_size} bytes)")
    
    print(f"\n✓ Sample Excel files created in: {os.path.abspath(folder_path)}")
    print(f"\nTo use these files:")
    print(f"1. Go to Data Integration tab")
    print(f"2. Create a new integration with folder path: {os.path.abspath(folder_path)}")
    print(f"3. Click 'Scan Folder' button")
    print(f"4. Files will be classified and appear in Trial Data Management tab")

if __name__ == "__main__":
    try:
        create_sample_excel_files()
    except ImportError:
        print("✗ openpyxl not installed. Installing...")
        import subprocess
        subprocess.check_call(["pip", "install", "openpyxl"])
        create_sample_excel_files()
    except Exception as e:
        print(f"✗ Error creating sample files: {e}")
